"""Tests unitaires pour le module d'intégration des achats Google Drive."""

import io
import json
from datetime import date
from pathlib import Path

import pytest

# ── Helpers pour créer des données de test ────────────────────────────────────

def _make_excel_bytes(purchase_cols: list[tuple[str, str, dict[str, float]]]) -> bytes:
    """
    Crée un fichier Excel de test en mémoire.

    purchase_cols : liste de (buyer, date_iso, {label: qty})
    Retourne les bytes du fichier .xlsx.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Ligne 0 : titre
    ws.cell(row=1, column=1, value="STOCK - ENTREES DES ACHATS")
    # Ligne 1 : màj
    ws.cell(row=2, column=1, value="màj le 28/04/2026")
    # Ligne 2 : marqueurs exemple — aucun dans nos tests
    # Ligne 3 : prénoms acheteurs (col C+)
    for i, (buyer, _, _) in enumerate(purchase_cols):
        ws.cell(row=4, column=3 + i, value=buyer)
    # Ligne 4 : dates d'achat (col C+)
    for i, (_, date_str, _) in enumerate(purchase_cols):
        ws.cell(row=5, column=3 + i, value=date_str)
    # Ligne 5 : vide
    # Ligne 6 : en-tête catégorie
    ws.cell(row=6, column=1, value="nb paquets")
    ws.cell(row=6, column=2, value="SOFTS")
    # Lignes 7+ : produits
    all_labels = sorted({lbl for _, _, items in purchase_cols for lbl in items})
    for row_offset, label in enumerate(all_labels):
        row = 7 + row_offset
        ws.cell(row=row, column=2, value=label)
        for i, (_, _, items) in enumerate(purchase_cols):
            qty = items.get(label)
            if qty is not None:
                ws.cell(row=row, column=3 + i, value=qty)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_raw_items(skus: list[str]) -> list:
    """Crée une liste minimale de raw_items pour les tests."""
    items = []
    for sku in skus:
        items.append({
            "stock_sku": sku,
            "label": sku,
            "enabled": True,
            "unit": "piece",
            "is_stock_reference": True,
            "consumption_per_sale": 1,
            "stock_state": {
                "stock_on_hand": 10.0,
                "stock_reserved": 0,
                "incoming_qty": 0,
                "incoming_eta": "",
                "last_inventory_date": "2026-01-01",
                "inventory_count_method": "manual",
                "stock_history": [],
            },
        })
    return items


# ── Tests : extract_file_id_from_url ─────────────────────────────────────────

class TestExtractFileIdFromUrl:
    """Tests de gdrive_loader.extract_file_id_from_url."""

    def setup_method(self):
        from stocks.gdrive_loader import extract_file_id_from_url
        self.extract = extract_file_id_from_url

    def test_spreadsheet_url(self):
        url = "https://docs.google.com/spreadsheets/d/1aBcDeFgHiJkLmNoPqRsTuVwXyZ01/edit"
        assert self.extract(url) == "1aBcDeFgHiJkLmNoPqRsTuVwXyZ01"

    def test_drive_file_url(self):
        url = "https://drive.google.com/file/d/1xYzAbCdEfGhIjKlMnOpQrStUv12/view"
        assert self.extract(url) == "1xYzAbCdEfGhIjKlMnOpQrStUv12"

    def test_raw_id(self):
        raw = "1aBcDeFgHiJkLmNoPqRsTuVwXyZ01"
        assert self.extract(raw) == raw

    def test_invalid_url_raises(self):
        with pytest.raises(ValueError):
            self.extract("https://example.com/not-a-drive-url")

    def test_short_string_raises(self):
        with pytest.raises(ValueError):
            self.extract("toocourt/path")


# ── Tests : parse_purchases_excel ─────────────────────────────────────────────

class TestParsePurchasesExcel:
    """Tests de update_stock_from_purchases.parse_purchases_excel."""

    def setup_method(self):
        from stocks.update_stock_from_purchases import parse_purchases_excel
        self.parse = parse_purchases_excel

    def test_single_purchase_column(self):
        xlsx = _make_excel_bytes([
            ("Alice", "2026-04-20", {"chips": 18.0, "coca": 5.0}),
        ])
        events = self.parse(xlsx)
        assert len(events) == 1
        assert events[0].buyer == "Alice"
        assert events[0].purchase_date == date(2026, 4, 20)
        labels = {item.excel_label for item in events[0].items}
        assert "chips" in labels
        assert "coca" in labels

    def test_multiple_purchase_columns(self):
        xlsx = _make_excel_bytes([
            ("Bob", "2026-04-10", {"chips": 10.0}),
            ("Carol", "2026-04-15", {"coca": 24.0}),
        ])
        events = self.parse(xlsx)
        assert len(events) == 2
        dates = {e.purchase_date for e in events}
        assert date(2026, 4, 10) in dates
        assert date(2026, 4, 15) in dates

    def test_exemple_columns_ignored(self):
        """Les colonnes marquées 'exemple' en ligne 2 doivent être ignorées."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="STOCK - ENTREES DES ACHATS")
        ws.cell(row=2, column=1, value="màj")
        ws.cell(row=3, column=3, value="exemple")   # colonne C = exemple
        ws.cell(row=3, column=4, value="")           # colonne D = pas exemple
        ws.cell(row=4, column=3, value="Corinne")
        ws.cell(row=4, column=4, value="Alice")
        ws.cell(row=5, column=3, value="2026-04-11")
        ws.cell(row=5, column=4, value="2026-04-20")
        ws.cell(row=6, column=1, value="nb paquets")  # ligne vide/unité
        # Produits à partir de la ligne 7 (rows[6:] dans le parser)
        ws.cell(row=7, column=2, value="chips")
        ws.cell(row=7, column=3, value=10.0)   # Corinne (ignorée)
        ws.cell(row=7, column=4, value=5.0)    # Alice (conservée)
        buf = io.BytesIO()
        wb.save(buf)

        events = self.parse(buf.getvalue())
        assert len(events) == 1
        assert events[0].buyer == "Alice"
        assert events[0].purchase_date == date(2026, 4, 20)

    def test_zero_qty_ignored(self):
        xlsx = _make_excel_bytes([
            ("Alice", "2026-04-20", {"chips": 0.0, "coca": 5.0}),
        ])
        events = self.parse(xlsx)
        assert len(events) == 1
        labels = [item.excel_label for item in events[0].items]
        assert "chips" not in labels
        assert "coca" in labels

    def test_category_headers_not_parsed_as_products(self):
        """Les lignes en MAJUSCULES (catégories) ne doivent pas être des produits."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="STOCK - ENTREES DES ACHATS")
        ws.cell(row=2, column=1, value="màj")
        ws.cell(row=4, column=3, value="Alice")
        ws.cell(row=5, column=3, value="2026-04-20")
        ws.cell(row=6, column=2, value="BOISSONS CHAUDES")  # catégorie
        ws.cell(row=6, column=3, value=99.0)
        ws.cell(row=7, column=2, value="coca")              # produit réel
        ws.cell(row=7, column=3, value=5.0)
        buf = io.BytesIO()
        wb.save(buf)

        events = self.parse(buf.getvalue())
        assert len(events) == 1
        labels = [item.excel_label for item in events[0].items]
        assert "BOISSONS CHAUDES" not in labels
        assert "coca" in labels

    def test_no_purchase_columns_returns_empty(self):
        """Un fichier avec les 6 lignes structurelles mais sans colonne d'achat retourne []."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="STOCK - ENTREES DES ACHATS")
        ws.cell(row=2, column=1, value="màj")
        # Ligne 3 : pas de marqueurs exemple
        # Ligne 4 : pas d'acheteurs (col C vide)
        ws.cell(row=5, column=2, value="date achat")  # col B uniquement
        ws.cell(row=6, column=1, value="nb paquets")
        ws.cell(row=7, column=2, value="chips")
        buf = io.BytesIO()
        wb.save(buf)
        events = self.parse(buf.getvalue())
        assert events == []


# ── Tests : load_purchase_mapping ────────────────────────────────────────────

class TestLoadPurchaseMapping:
    """Tests de update_stock_from_purchases.load_purchase_mapping."""

    def setup_method(self):
        from stocks.update_stock_from_purchases import load_purchase_mapping
        self.load = load_purchase_mapping

    def test_basic_mapping_loaded(self, tmp_path):
        data = {
            "products": [
                {"excel_label": "chips", "stock_sku": "chips_sku", "qty_multiplier": 1},
                {"excel_label": "Café en grains 1 kg", "stock_sku": "cafe", "qty_multiplier": 1000},
            ]
        }
        p = tmp_path / "mapping.json"
        p.write_text(json.dumps(data), encoding="utf-8")
        mapping = self.load(p)
        assert "chips" in mapping
        assert mapping["chips"] == ("chips_sku", 1.0)

    def test_label_normalized(self, tmp_path):
        """Les labels sont normalisés (minuscules, sans accents)."""
        data = {"products": [
            {"excel_label": "Café en grains 1 kg", "stock_sku": "cafe", "qty_multiplier": 1000},
        ]}
        p = tmp_path / "mapping.json"
        p.write_text(json.dumps(data), encoding="utf-8")
        mapping = self.load(p)
        assert "cafe en grains 1 kg" in mapping
        assert mapping["cafe en grains 1 kg"] == ("cafe", 1000.0)

    def test_multiplier_applied(self, tmp_path):
        data = {"products": [
            {"excel_label": "the boite", "stock_sku": "the", "qty_multiplier": 25},
        ]}
        p = tmp_path / "mapping.json"
        p.write_text(json.dumps(data), encoding="utf-8")
        mapping = self.load(p)
        _, mult = mapping["the boite"]
        assert mult == 25.0


# ── Tests : find_already_processed_dates ─────────────────────────────────────

class TestFindAlreadyProcessedDates:
    """Tests de update_stock_from_purchases.find_already_processed_dates."""

    def setup_method(self):
        from stocks.update_stock_from_purchases import find_already_processed_dates
        self.find = find_already_processed_dates

    def test_no_history(self):
        items = _make_raw_items(["chips"])
        assert self.find(items) == set()

    def test_purchase_dates_detected(self):
        items = _make_raw_items(["chips"])
        items[0]["stock_state"]["stock_history"] = [
            {"type": "purchase", "date": "2026-04-20", "qty_added": 18.0},
            {"type": "auto_refresh", "from_date": "2026-04-16", "to_date": "2026-04-20"},
        ]
        processed = self.find(items)
        assert date(2026, 4, 20) in processed
        assert len(processed) == 1

    def test_multiple_items_aggregated(self):
        items = _make_raw_items(["chips", "coca"])
        items[0]["stock_state"]["stock_history"] = [
            {"type": "purchase", "date": "2026-04-10"},
        ]
        items[1]["stock_state"]["stock_history"] = [
            {"type": "purchase", "date": "2026-04-15"},
        ]
        processed = self.find(items)
        assert date(2026, 4, 10) in processed
        assert date(2026, 4, 15) in processed

    def test_invalid_date_ignored(self):
        items = _make_raw_items(["chips"])
        items[0]["stock_state"]["stock_history"] = [
            {"type": "purchase", "date": "not-a-date"},
        ]
        assert self.find(items) == set()


# ── Tests : apply_purchases_to_stock ─────────────────────────────────────────

class TestApplyPurchasesToStock:
    """Tests de update_stock_from_purchases.apply_purchases_to_stock."""

    def setup_method(self):
        from stocks.update_stock_from_purchases import apply_purchases_to_stock, PurchaseEvent, PurchaseItem
        self.apply = apply_purchases_to_stock
        self.PurchaseEvent = PurchaseEvent
        self.PurchaseItem = PurchaseItem

    def _event(self, d: str, buyer: str, items: dict[str, float]):
        return self.PurchaseEvent(
            purchase_date=date.fromisoformat(d),
            buyer=buyer,
            items=[self.PurchaseItem(excel_label=k, qty=v) for k, v in items.items()],
        )

    def test_basic_stock_increase(self):
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        event = self._event("2026-04-20", "Alice", {"chips": 18.0})

        items, successes, warnings = self.apply(items, [event], mapping, set())

        assert not warnings
        assert len(successes) == 1
        assert items[0]["stock_state"]["stock_on_hand"] == 28.0  # 10 + 18

    def test_unit_multiplier_applied(self):
        items = _make_raw_items(["cafe"])
        mapping = {"cafe 1 kg": ("cafe", 1000.0)}
        event = self._event("2026-04-20", "Bob", {"cafe 1 kg": 2.0})

        items, successes, warnings = self.apply(items, [event], mapping, set())

        assert not warnings
        assert items[0]["stock_state"]["stock_on_hand"] == 10.0 + 2000.0

    def test_deduplication_skips_already_processed(self):
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        event = self._event("2026-04-20", "Alice", {"chips": 18.0})
        already = {date(2026, 4, 20)}

        items, successes, warnings = self.apply(items, [event], mapping, already)

        assert successes == []
        assert items[0]["stock_state"]["stock_on_hand"] == 10.0  # inchangé

    def test_history_entry_added(self):
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        event = self._event("2026-04-20", "Alice", {"chips": 18.0})

        items, _, _ = self.apply(items, [event], mapping, set())

        history = items[0]["stock_state"]["stock_history"]
        assert len(history) == 1
        entry = history[0]
        assert entry["type"] == "purchase"
        assert entry["date"] == "2026-04-20"
        assert entry["buyer"] == "Alice"
        assert entry["qty_added"] == 18.0
        assert entry["source"] == "gdrive_excel"

    def test_dry_run_no_modification(self):
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        event = self._event("2026-04-20", "Alice", {"chips": 18.0})

        items, successes, _ = self.apply(items, [event], mapping, set(), dry_run=True)

        assert len(successes) == 1
        assert items[0]["stock_state"]["stock_on_hand"] == 10.0  # non modifié
        assert items[0]["stock_state"]["stock_history"] == []    # pas d'entrée

    def test_unknown_label_produces_warning(self):
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        event = self._event("2026-04-20", "Alice", {"produit_inconnu": 5.0})

        _, successes, warnings = self.apply(items, [event], mapping, set())

        assert successes == []
        assert len(warnings) == 1
        assert "produit_inconnu" in warnings[0]

    def test_unknown_sku_produces_warning(self):
        items = _make_raw_items(["chips"])
        mapping = {"soda": ("sku_inexistant", 1.0)}
        event = self._event("2026-04-20", "Alice", {"soda": 5.0})

        _, successes, warnings = self.apply(items, [event], mapping, set())

        assert successes == []
        assert len(warnings) == 1
        assert "sku_inexistant" in warnings[0]

    def test_events_sorted_by_date(self):
        """Les événements doivent être traités dans l'ordre chronologique."""
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        events = [
            self._event("2026-04-25", "Carol", {"chips": 5.0}),
            self._event("2026-04-10", "Alice", {"chips": 3.0}),
        ]
        items, _, _ = self.apply(items, events, mapping, set())
        # Stock final : 10 + 3 + 5 = 18
        assert items[0]["stock_state"]["stock_on_hand"] == 18.0
        history = items[0]["stock_state"]["stock_history"]
        assert history[0]["date"] == "2026-04-10"
        assert history[1]["date"] == "2026-04-25"

    def test_already_processed_set_updated_after_apply(self):
        """La date doit être ajoutée à already_processed après traitement."""
        items = _make_raw_items(["chips"])
        mapping = {"chips": ("chips", 1.0)}
        event = self._event("2026-04-20", "Alice", {"chips": 5.0})
        already = set()

        self.apply(items, [event], mapping, already)

        assert date(2026, 4, 20) in already


# ── Tests : sauvegarde atomique ───────────────────────────────────────────────

class TestSaveStockItems:
    """Tests de update_stock_from_purchases._save_stock_items."""

    def setup_method(self):
        from stocks.update_stock_from_purchases import _save_stock_items
        self.save = _save_stock_items

    def test_atomic_write(self, tmp_path):
        path = tmp_path / "stock_items.json"
        data = [{"stock_sku": "chips", "stock_state": {"stock_on_hand": 42}}]
        self.save(path, data)
        assert path.exists()
        loaded = json.loads(path.read_text(encoding="utf-8"))
        assert loaded[0]["stock_state"]["stock_on_hand"] == 42

    def test_no_tmp_file_left(self, tmp_path):
        path = tmp_path / "stock_items.json"
        self.save(path, [{"a": 1}])
        tmp = Path(str(path) + ".tmp")
        assert not tmp.exists()

    def test_overwrites_existing(self, tmp_path):
        path = tmp_path / "stock_items.json"
        self.save(path, [{"v": 1}])
        self.save(path, [{"v": 2}])
        loaded = json.loads(path.read_text(encoding="utf-8"))
        assert loaded[0]["v"] == 2

    def test_roundtrip_preserves_data(self, tmp_path):
        from stocks.update_stock_from_purchases import _load_stock_items_raw
        path = tmp_path / "stock_items.json"
        original = _make_raw_items(["chips", "coca"])
        self.save(path, original)
        reloaded = _load_stock_items_raw(path)
        assert len(reloaded) == 2
        assert reloaded[0]["stock_sku"] == "chips"
