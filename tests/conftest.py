"""Fixtures partagées entre tous les fichiers de tests."""
import io
import os

import pytest

# Variables d'environnement requises par les modules avant leur import
os.environ.setdefault("SUMUP_API_KEY", "test_api_key_fixture")
os.environ.setdefault("PAHEKO_BASE_URL", "http://localhost:9999")
os.environ.setdefault("PAHEKO_API_USER", "test_user")
os.environ.setdefault("PAHEKO_API_PASSWORD", "test_pass")


# ── Fixtures stock items ───────────────────────────────────────────────────────

@pytest.fixture
def raw_items():
    """Liste minimale d'articles de stock valides avec stock_state."""
    return [
        {
            "stock_sku": "chips",
            "label": "Chips",
            "stock_label": "Chips",
            "enabled": True,
            "unit": "piece",
            "category": "snacking",
            "is_stock_reference": True,
            "consumption_per_sale": 1,
            "sumup_match": {"name": "Chips", "variant": ""},
            "stock_state": {
                "stock_on_hand": 10.0,
                "stock_reserved": 0,
                "incoming_qty": 0,
                "incoming_eta": "",
                "last_inventory_date": "2026-01-01",
                "inventory_count_method": "manual",
                "stock_history": [],
            },
        },
        {
            "stock_sku": "coca",
            "label": "Coca-Cola",
            "stock_label": "Coca-Cola",
            "enabled": True,
            "unit": "piece",
            "category": "soft",
            "is_stock_reference": True,
            "consumption_per_sale": 1,
            "sumup_match": {"name": "Jus & Sodas", "variant": "Coca Cola classique"},
            "stock_state": {
                "stock_on_hand": 24.0,
                "stock_reserved": 0,
                "incoming_qty": 0,
                "incoming_eta": "",
                "last_inventory_date": "2026-01-01",
                "inventory_count_method": "manual",
                "stock_history": [],
            },
        },
    ]


@pytest.fixture
def sample_mapping():
    """Mapping label normalisé → (sku, multiplicateur) simple."""
    return {
        "chips": ("chips", 1.0),
        "coca": ("coca", 1.0),
        "cafe 1 kg": ("cafe", 1000.0),
    }


@pytest.fixture
def excel_bytes_factory():
    """Factory : retourne une fonction créant un xlsx en mémoire."""

    def _make(purchase_cols: list[tuple[str, str, dict]]) -> bytes:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="STOCK - ENTREES DES ACHATS")
        ws.cell(row=2, column=1, value="màj le 28/04/2026")
        for i, (buyer, _, _) in enumerate(purchase_cols):
            ws.cell(row=4, column=3 + i, value=buyer)
        for i, (_, date_str, _) in enumerate(purchase_cols):
            ws.cell(row=5, column=3 + i, value=date_str)
        ws.cell(row=6, column=1, value="nb paquets")
        ws.cell(row=6, column=2, value="SOFTS")
        all_labels = sorted({lbl for _, _, items in purchase_cols for lbl in items})
        for row_offset, label in enumerate(all_labels):
            row = 7 + row_offset
            ws.cell(row=row, column=2, value=label)
            for i, (_, _, items) in enumerate(purchase_cols):
                qty = items.get(label)
                if qty is not None:
                    ws.cell(row=row, column=3 + i, value=qty)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    return _make


@pytest.fixture
def raw_catalog_items():
    """Articles bruts compatibles sumup_statistics.Catalog."""
    return [
        {
            "stock_sku": "chips",
            "label": "Chips",
            "stock_label": "Chips",
            "category": "snacking",
            "unit": "piece",
            "enabled": True,
            "is_stock_reference": True,
            "consumption_per_sale": 1,
            "sumup_match": {"name": "Chips", "variant": ""},
            "stock_state": {"stock_on_hand": 10},
        },
        {
            "stock_sku": "coca",
            "label": "Coca-Cola",
            "stock_label": "Coca-Cola",
            "category": "soft",
            "unit": "piece",
            "enabled": True,
            "is_stock_reference": True,
            "consumption_per_sale": 1,
            "sumup_match": {"name": "Jus & Sodas", "variant": "Coca Cola classique"},
            "stock_state": {"stock_on_hand": 24},
        },
        {
            "stock_sku": "disabled_item",
            "label": "Désactivé",
            "category": "autres",
            "unit": "piece",
            "enabled": False,
            "is_stock_reference": True,
            "consumption_per_sale": 1,
            "sumup_match": {"name": "Disabled", "variant": ""},
        },
    ]
